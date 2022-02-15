# multiplication_table_maker.py

from pathlib import Path

import openpyxl
from openpyxl.styles import Font


def multplication_table(n: int, name: str):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f"{n}x{n} multiplication"

    bold = Font(bold=True)

    for i in range(1, n + 1):
        sheet.cell(row=1, column=i + 1, value=i).font = bold  # type: ignore
        sheet.cell(row=i + 1, column=1, value=i).font = bold  # type: ignore

        for k in range(1, n + 1):
            sheet.cell(row=i + 1, column=k + 1, value=i * k)

    file_name = name + ".xlsx"
    path = Path(__file__).parent.resolve() / file_name
    wb.save(str(path))


if __name__ == "__main__":
    number = 6
    name = f"multiplication_by_{number}"
    multplication_table(number, name)
