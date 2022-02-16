# spreadsheet_cell_inverter.py

from pathlib import Path

import openpyxl


def invert_spreadsheet_cells(file_name: str) -> None:
    directory = Path(__file__).parent.resolve()
    wb_path = directory / file_name

    wb = openpyxl.load_workbook(str(wb_path))
    sheet = wb.active

    modified_wb = openpyxl.Workbook()
    modified_sheet = modified_wb.active

    for i in range(1, sheet.max_row + 1):
        for k in range(1, sheet.max_column + 1):
            old_value = sheet.cell(row=i, column=k).value
            modified_sheet.cell(row=k, column=i, value=old_value)

    modified_name = file_name.split(".")[0] + "_invert.xlsx"
    modified_path = directory / modified_name
    modified_wb.save(str(modified_path))

if __name__ == "__main__":
    file_name = "shopping_list.xlsx"
    invert_spreadsheet_cells(file_name)
