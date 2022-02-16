# spreadsheet_to_text_files.py

from pathlib import Path

import openpyxl


def xlsx_to_txt(file_name: str) -> None:
    directory = Path(__file__).parent.resolve()
    path = directory / file_name

    wb = openpyxl.load_workbook(str(path))
    sheet = wb.active

    files_text = []

    for column in range(1, sheet.max_column + 1):
        file_text = []

        for row in range(1, sheet.max_row + 1):
            line = sheet.cell(row=row, column=column).value
            if line != None:
                file_text.append(line)

        files_text.append(file_text)

    print(files_text)

    text_file_name = 1

    for new_file in files_text:
        file_path = directory / f"File_{text_file_name}.txt"
        file = open(file_path, "w")

        for line in new_file:
            file.write(line)

        file.close()
        text_file_name += 1


if __name__ == "__main__":
    file_name = "xlsx_to_txt.xlsx"
    xlsx_to_txt(file_name)
