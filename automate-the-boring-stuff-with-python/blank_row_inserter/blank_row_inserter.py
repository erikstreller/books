# blank_row_inserter.py

from pathlib import Path

import openpyxl


def insert_blank_rows(file_name: str, n: int, m: int) -> None:
    directory = Path(__file__).parent.resolve()
    wb_path = directory / file_name

    wb = openpyxl.load_workbook(str(wb_path))
    sheet = wb.active

    modified_wb = openpyxl.Workbook()
    modified_sheet = modified_wb.active

    for i in range(1, sheet.max_row + 1):
        for k in range(1, sheet.max_column + 1):
            old_value = sheet.cell(row=i, column=k).value

            if i < n:
                modified_sheet.cell(row=i, column=k, value=old_value)
            else:
                modified_sheet.cell(row=i + m, column=k, value=old_value)

    modified_name = file_name.split(".")[0] + "_blank.xlsx"
    modified_path = directory / modified_name
    modified_wb.save(str(modified_path))


if __name__ == "__main__":
    file_name = "random_numbers.xlsx"
    n = 3
    m = 2
    insert_blank_rows(file_name, n, m)
